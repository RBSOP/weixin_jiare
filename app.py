"""Web 服务 - 数据展示页"""
import io
import json
import logging
from datetime import datetime
from pathlib import Path

from flask import Flask, Response, jsonify, render_template, request, send_file

# 关闭 Werkzeug 的 HTTP 请求日志，仅保留核心业务日志
logging.getLogger("werkzeug").disabled = True

from config import LOG_FILE, SCREENSHOT_DIR, UI_CONFIG_FILE
from db import clear_all_data, delete_account, list_accounts
from db_query import query_orders_with_relations
from export_xlsx import (
    SCREENSHOT_COLS,
    SCREENSHOT_LABELS,
    add_screenshot_image,
    format_export_value,
    get_category_colors,
    get_screenshot_path,
)
from pending_action import write as write_pending_action

app = Flask(__name__, template_folder="templates")
app.config["JSON_AS_ASCII"] = False

_ORDERS_COL_MAP = {
    "orders_nick_name": "加热主播",
    "orders_create_time": "创建时间",
    "orders_order_name": "名称",
    "promotion_id": "编号",
    "orders_target": "目标",
    "orders_status": "状态",
    "orders_budget": "预算",
    "orders_bid_roi": "出价/目标ROI",
    "orders_duration": "实际加热时长",
    "orders_cost_yuan": "消耗",
    "orders_actual_roi": "ROI",
}
_CATEGORY_PREFIXES = [
    ("orders_", "订单基本信息"),
    ("detail_直播间加热效果", "直播间加热效果"),
    ("detail_电商加热效果", "电商加热效果(详情)"),
    ("create_config_", "再来一单配置"),
    ("detail_加热信息", "加热信息"),
    ("detail_十分钟级数据", "十分钟级数据"),
    ("detail_", "其他"),
    ("ecommerce_", "电商加热效果"),
    ("people_funnel_", "人群漏斗"),
    ("people_feature_观众商品偏好_", "观众商品偏好"),
    ("people_feature_八类人群占比_", "八类人群占比"),
    ("people_feature_性别分布_", "性别分布"),
    ("people_feature_年龄分布_", "年龄分布"),
    ("people_feature_地域分布_", "地域分布"),
    ("screenshot_", "截图"),
]

# 加热信息列顺序（与页面一致，观众兴趣数值来自再来一单配置）
_HEATING_INFO_COL_ORDER = [
    "detail_加热信息_订单名称",
    "detail_加热信息_编号",
    "detail_加热信息_加热目标",
    "detail_加热信息_状态",
    "detail_加热信息_加热预算",
    "detail_加热信息_预计时长",
    "detail_加热信息_下单时间",
    "detail_加热信息_开始时间",
    "detail_加热信息_结束时间",
    "detail_加热信息_实际加热时长",
    "detail_加热信息_加热出价",
    "detail_加热信息_放量模式",
    "detail_加热信息_人群定向_观众性别",
    "detail_加热信息_人群定向_根据粉丝层推荐",
    "detail_加热信息_人群定向_根据名单推荐",
    "detail_加热信息_人群定向_观众城市",
    "detail_加热信息_人群定向_观众年龄",
    "detail_加热信息_人群定向_观众设备",
    "detail_加热信息_人群定向_观众兴趣",
]

_HEATING_INFO_DISPLAY_NAMES = {
    "detail_加热信息_订单名称": "名称",
    "detail_加热信息_人群定向_观众性别": "观众性别",
    "detail_加热信息_人群定向_根据粉丝层推荐": "根据粉丝层推荐",
    "detail_加热信息_人群定向_根据名单推荐": "根据名单推荐",
    "detail_加热信息_人群定向_观众城市": "观众城市",
    "detail_加热信息_人群定向_观众年龄": "观众年龄",
    "detail_加热信息_人群定向_观众设备": "观众设备",
    "detail_加热信息_人群定向_观众兴趣": "观众兴趣",
}

_TEN_MIN_DISPLAY_NAMES = {
    "detail_十分钟级数据_时间": "时间",
    "detail_十分钟级数据_直播间消耗": "直播间消耗",
    "detail_十分钟级数据_直播间曝光人数": "直播间曝光人数",
    "detail_十分钟级数据_直播间观看人数": "直播间观看人数",
    "detail_十分钟级数据_直播间点赞次数": "直播间点赞次数",
    "detail_十分钟级数据_直播间评论次数": "直播间评论次数",
    "detail_十分钟级数据_直播间新增粉丝数": "直播间新增粉丝数",
}

# 直播间加热效果列顺序（与 DOM 一致）
_EFFECT_SUMMARY_COL_ORDER = [
    "detail_直播间加热效果_消耗总金额",
    "detail_直播间加热效果_曝光总人数",
    "detail_直播间加热效果_进入总人数",
    "detail_直播间加热效果_点赞总次数",
    "detail_直播间加热效果_评论总次数",
    "detail_直播间加热效果_新增总关注",
]

# 电商加热效果(详情页)列顺序（与 DOM 一致）
_ECOMMERCE_EFFECT_COL_ORDER = [
    "detail_电商加热效果_总成交ROI",
    "detail_电商加热效果_成交GMV",
    "detail_电商加热效果_商品点击人数",
    "detail_电商加热效果_商品点击次数",
    "detail_电商加热效果_下单订单数",
    "detail_电商加热效果_下单GMV",
    "detail_电商加热效果_成交订单数",
]

# 人群分布各块列顺序（一级表头分块，二级表头 观众商品偏好1、2、3...）
_PEOPLE_FEATURE_BLOCKS = [
    ("观众商品偏好", [f"people_feature_观众商品偏好_{i}" for i in range(1, 9)]),
    ("八类人群占比", [f"people_feature_八类人群占比_{i}" for i in range(1, 9)]),
    ("性别分布", [f"people_feature_性别分布_{i}" for i in range(1, 3)]),
    ("年龄分布", [f"people_feature_年龄分布_{i}" for i in range(1, 9)]),
    ("地域分布", [f"people_feature_地域分布_{i}" for i in range(1, 9)]),
]
_PEOPLE_FEATURE_BLOCK_ORDER = [c for _, cols in _PEOPLE_FEATURE_BLOCKS for c in cols]

# 人群漏斗列顺序（与页面漏斗图一致：人数自上而下 + 转化率）
_PEOPLE_FUNNEL_COL_ORDER = [
    "people_funnel_直播间曝光人数",
    "people_funnel_直播间观看人数",
    "people_funnel_商品曝光人数",
    "people_funnel_商品点击人数",
    "people_funnel_总下单人数",
    "people_funnel_总成交人数",
    "people_funnel_直播间点击率",
    "people_funnel_商品曝光率",
    "people_funnel_商品点击率",
    "people_funnel_点击下单率",
    "people_funnel_下单成交率",
    "people_funnel_总成交转化率",
]

# 电商加热效果列顺序：12 项自选指标 + 9 项数据明细
_ECOMMERCE_COL_ORDER = [
    "ecommerce_总消耗",
    "ecommerce_直接成交金额",
    "ecommerce_直接成交订单数",
    "ecommerce_直接成交ROI",
    "ecommerce_直播间曝光人数",
    "ecommerce_短视频曝光次数",
    "ecommerce_直播间观看人数",
    "ecommerce_直播间商品点击率",
    "ecommerce_直播间平均千次展示费用",
    "ecommerce_总成交ROI",
    "ecommerce_间接成交金额",
    "ecommerce_直播间观看人次",
    "ecommerce_加热主播与创建时间",
    "ecommerce_名称与编号",
    "ecommerce_加热开始时间",
    "ecommerce_加热结束时间",
    "ecommerce_实际加热时长",
    "ecommerce_直播间进入率（人数）",
    "ecommerce_GPM",
    "ecommerce_直播间转化率",
    "ecommerce_直播间转化成本",
]

# 十分钟级数据列顺序
_TEN_MIN_COL_ORDER = [
    "detail_十分钟级数据_时间",
    "detail_十分钟级数据_直播间消耗",
    "detail_十分钟级数据_直播间曝光人数",
    "detail_十分钟级数据_直播间观看人数",
    "detail_十分钟级数据_直播间点赞次数",
    "detail_十分钟级数据_直播间评论次数",
    "detail_十分钟级数据_直播间新增粉丝数",
]

# 再来一单配置列顺序（与原页面一致：选择加热类型→选择加热对象→选择加热方案→预算与时间→人群定向→其他）
_CREATE_CONFIG_COL_ORDER = [
    "create_config_选择加热类型_加热对象类型",
    "create_config_选择加热类型_加热订单类型",
    "create_config_选择加热对象_主播昵称",
    "create_config_选择加热方案_预计带来商品成交金额",
    "create_config_选择加热方案_基础信息_订单名称",
    "create_config_选择加热方案_基础信息_加热方式",
    "create_config_选择加热方案_基础信息_放量模式",
    "create_config_选择加热方案_基础信息_优先提升目标",
    "create_config_选择加热方案_基础信息_成交ROI",
    "create_config_选择加热方案_基础信息_加热素材",
    "create_config_预算与时间_订单预算",
    "create_config_预算与时间_加热时长",
    "create_config_人群定向_定向类型",
    "create_config_人群定向_观众性别",
    "create_config_人群定向_根据粉丝层推荐",
    "create_config_人群定向_根据名单推荐",
    "create_config_人群定向_观众年龄",
    "create_config_人群定向_观众设备",
    "create_config_人群定向_观众城市",
    "create_config_人群定向_观众兴趣",
    "create_config_其他_其他支付方式",
]


_CREATE_CONFIG_DISPLAY_NAMES = {
    "create_config_人群定向_观众年龄": "观众年龄",
}


def _col_display_name(col: str) -> str:
    if col in _ORDERS_COL_MAP:
        return _ORDERS_COL_MAP[col]
    if col in SCREENSHOT_LABELS:
        return SCREENSHOT_LABELS[col]
    if col in _CREATE_CONFIG_DISPLAY_NAMES:
        return _CREATE_CONFIG_DISPLAY_NAMES[col]
    if col in _HEATING_INFO_DISPLAY_NAMES:
        return _HEATING_INFO_DISPLAY_NAMES[col]
    if col.startswith("detail_加热信息"):
        suffix = col[len("detail_加热信息_"):]
        if suffix.startswith("人群定向_"):
            suffix = suffix[len("人群定向_"):]
        return suffix.replace("_", " ")
    if col in _TEN_MIN_DISPLAY_NAMES:
        return _TEN_MIN_DISPLAY_NAMES[col]
    if col.startswith("detail_直播间加热效果_"):
        return col[len("detail_直播间加热效果_"):].replace("_", " ")
    if col.startswith("detail_电商加热效果_"):
        return col[len("detail_电商加热效果_"):].replace("_", " ")
    if col.startswith("people_feature_"):
        parts = col.split("_", 3)
        if len(parts) >= 4:
            return f"{parts[2]}{parts[3]}"
    for prefix, _ in _CATEGORY_PREFIXES:
        if col.startswith(prefix):
            return col[len(prefix):].replace("_", " ")
    return col


def _is_other_category(col: str) -> bool:
    """检查列是否属于「其他」分类（不展示、不导出）"""
    if col == "promotion_id" or col.startswith("orders_"):
        return False
    if col in SCREENSHOT_COLS:
        return False
    if col.startswith("create_config_"):
        return False
    if col.startswith("detail_加热信息"):
        return False
    if col.startswith("detail_直播间加热效果"):
        return False
    if col.startswith("detail_电商加热效果"):
        return False
    if col.startswith("detail_十分钟级数据"):
        return False
    if col.startswith("ecommerce_"):
        return False
    if col.startswith("people_funnel_"):
        return False
    if col.startswith("people_feature_"):
        return False
    if col.startswith("detail_"):
        return True
    return True


def _get_col_category_map(cols: list[str]) -> dict[str, str]:
    m = {}
    for c in cols:
        if c == "promotion_id" or c.startswith("orders_"):
            m[c] = "订单基本信息"
            continue
        if c in SCREENSHOT_COLS:
            m[c] = "截图"
            continue
        if c.startswith("detail_加热信息"):
            m[c] = "加热信息"
            continue
        matched = False
        for prefix, label in _CATEGORY_PREFIXES:
            if c.startswith(prefix):
                m[c] = label
                matched = True
                break
        if not matched:
            m[c] = "其他"
    return m


def _sort_export_cols(cols: list[str]) -> list[str]:
    cat_order = [p for p, _ in _CATEGORY_PREFIXES] + [""]
    priority = list(_ORDERS_COL_MAP.keys())
    create_order_map = {k: i for i, k in enumerate(_CREATE_CONFIG_COL_ORDER)}
    heating_info_map = {k: i for i, k in enumerate(_HEATING_INFO_COL_ORDER)}
    effect_summary_map = {k: i for i, k in enumerate(_EFFECT_SUMMARY_COL_ORDER)}
    ecommerce_effect_map = {k: i for i, k in enumerate(_ECOMMERCE_EFFECT_COL_ORDER)}
    ten_min_map = {k: i for i, k in enumerate(_TEN_MIN_COL_ORDER)}
    ecommerce_order_map = {k: i for i, k in enumerate(_ECOMMERCE_COL_ORDER)}
    people_funnel_order_map = {k: i for i, k in enumerate(_PEOPLE_FUNNEL_COL_ORDER)}
    people_feature_order_map = {k: i for i, k in enumerate(_PEOPLE_FEATURE_BLOCK_ORDER)}
    screenshot_order_map = {k: i for i, k in enumerate(SCREENSHOT_COLS)}

    def sort_key(c):
        cat_idx = len(cat_order)
        if c == "promotion_id" or c.startswith("orders_"):
            cat_idx = 0
        elif c in SCREENSHOT_COLS:
            cat_idx = _CATEGORY_PREFIXES.index(("screenshot_", "截图"))
        else:
            for i, (prefix, _) in enumerate(_CATEGORY_PREFIXES):
                if c.startswith(prefix):
                    cat_idx = i
                    break
        pri = priority.index(c) if c in priority else 999
        create_idx = create_order_map.get(c, 9999) if c.startswith("create_config_") else 9999
        heating_idx = heating_info_map.get(c, 9999) if c.startswith("detail_加热信息") else 9999
        effect_idx = effect_summary_map.get(c, 9999) if c.startswith("detail_直播间加热效果") else 9999
        ecommerce_effect_idx = ecommerce_effect_map.get(c, 9999) if c.startswith("detail_电商加热效果") else 9999
        ten_idx = ten_min_map.get(c, 9999) if c.startswith("detail_十分钟级数据") else 9999
        ecommerce_idx = ecommerce_order_map.get(c, 9999) if c.startswith("ecommerce_") else 0
        funnel_idx = people_funnel_order_map.get(c, 9999) if c.startswith("people_funnel_") else 0
        feat_idx = people_feature_order_map.get(c, 9999) if c.startswith("people_feature_") else 9999
        screenshot_idx = screenshot_order_map.get(c, 9999) if c in SCREENSHOT_COLS else 9999
        return (
            cat_idx,
            create_idx if c.startswith("create_config_") else 0,
            heating_idx if c.startswith("detail_加热信息") else 0,
            effect_idx if c.startswith("detail_直播间加热效果") else 0,
            ecommerce_effect_idx if c.startswith("detail_电商加热效果") else 0,
            ten_idx if c.startswith("detail_十分钟级数据") else 0,
            ecommerce_idx if c.startswith("ecommerce_") else 0,
            funnel_idx if c.startswith("people_funnel_") else 0,
            feat_idx if c.startswith("people_feature_") else 9999,
            screenshot_idx if c in SCREENSHOT_COLS else 9999,
            pri,
            c,
        )

    return sorted(cols, key=sort_key)

DEFAULT_UI_CONFIG = {"mode": "prod", "resolution": "1280x800"}


@app.route("/")
def index():
    """数据展示页"""
    return render_template("index.html")


@app.route("/api/data")
def api_data():
    """分页查询数据"""
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 20))
    cost_min = request.args.get("cost_min", type=float)
    promotion_id = request.args.get("promotion_id", "").strip()
    nick_name = request.args.get("nick_name", "").strip()
    order_name = request.args.get("order_name", "").strip()
    sort_by = request.args.get("sort_by", "create_time")
    sort_order = request.args.get("sort_order", "desc")

    rows, total = query_orders_with_relations(
        cost_min=cost_min,
        promotion_id=promotion_id or None,
        nick_name=nick_name or None,
        order_name=order_name or None,
        page=page,
        page_size=page_size,
        sort_by=sort_by,
        sort_order=sort_order,
    )
    return jsonify({"rows": rows, "total": total, "page": page, "page_size": page_size})


_SCREENSHOT_PAGE_TYPES = ("create_config", "order_detail", "ecommerce", "people")


@app.route("/api/screenshot/<path:promotion_id>/<page_type>")
def api_screenshot(promotion_id: str, page_type: str):
    """返回指定订单的页面截图，默认不加载，用户点击后请求"""
    if page_type not in _SCREENSHOT_PAGE_TYPES:
        return jsonify({"error": "未知页面类型"}), 404
    safe_pid = promotion_id.replace("/", "_").replace("\\", "_")
    path = Path(SCREENSHOT_DIR) / safe_pid / f"{page_type}.png"
    if not path.exists():
        return jsonify({"error": "暂无截图"}), 404
    return send_file(str(path), mimetype="image/png")


@app.route("/api/export")
def api_export():
    """导出 xlsx（按当前筛选条件）"""
    cost_min = request.args.get("cost_min", type=float)
    promotion_id = request.args.get("promotion_id", "").strip()
    nick_name = request.args.get("nick_name", "").strip()
    order_name = request.args.get("order_name", "").strip()
    sort_by = request.args.get("sort_by", "create_time")
    sort_order = request.args.get("sort_order", "desc")

    rows, _ = query_orders_with_relations(
        cost_min=cost_min,
        promotion_id=promotion_id or None,
        nick_name=nick_name or None,
        order_name=order_name or None,
        page=1,
        page_size=10000,
        sort_by=sort_by.replace("orders_", ""),
        sort_order=sort_order,
    )
    if not rows:
        return jsonify({"error": "无数据可导出"}), 400

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "未安装 openpyxl"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "采集数据"

    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())
    hide = {"orders_account_id", "orders_cost", "orders_created_at", "orders_promotion_id"}
    heating_info_allowed = set(_HEATING_INFO_COL_ORDER)
    ecommerce_allowed = set(_ECOMMERCE_COL_ORDER)

    def _export_col_filter(k: str) -> bool:
        if k in hide:
            return False
        if k.startswith("detail_加热信息") and k not in heating_info_allowed:
            return False
        if k.startswith("ecommerce_") and k not in ecommerce_allowed:
            return False
        if k in {f"people_feature_性别分布_{i}" for i in range(3, 9)}:
            return False
        return not _is_other_category(k)

    cols = [k for k in all_keys if _export_col_filter(k)]
    for c in SCREENSHOT_COLS:
        if c not in cols:
            cols.append(c)
    cols = _sort_export_cols(cols)

    ten_min_cols = {c for c in cols if c.startswith("detail_十分钟级数据_")}

    category_map = _get_col_category_map(cols)
    groups = []
    cur_cat = None
    for c in cols:
        cat = category_map.get(c, "其他")
        if cat != cur_cat:
            groups.append({"label": cat, "start": cols.index(c) + 1, "count": 1})
            cur_cat = cat
        else:
            groups[-1]["count"] += 1

    from openpyxl.styles import Font, Alignment, PatternFill

    screenshot_dir = Path(SCREENSHOT_DIR)

    for g in groups:
        start_col = g["start"]
        end_col = start_col + g["count"] - 1
        cat = g["label"]
        dark, light = get_category_colors(cat)
        header_fill = PatternFill(start_color=dark, end_color=dark, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        if g["count"] > 1:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(1, start_col, cat)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ci, k in enumerate(cols, 1):
        cat = category_map.get(k, "其他")
        _, light = get_category_colors(cat)
        cell = ws.cell(2, ci, _col_display_name(k))
        cell.fill = PatternFill(start_color=light, end_color=light, fill_type="solid")
        cell.font = Font(bold=True, size=11)

    row_idx = 3
    promotion_groups = []
    cur_pid = None
    screenshot_images = []
    for r in rows:
        pid = r.get("promotion_id")
        is_first_of_group = pid != cur_pid
        if pid != cur_pid:
            promotion_groups.append({"pid": pid, "start": row_idx, "count": 1})
            cur_pid = pid
        else:
            promotion_groups[-1]["count"] += 1
        for ci, k in enumerate(cols, 1):
            if k in SCREENSHOT_COLS:
                page_type = k.replace("screenshot_", "")
                path = get_screenshot_path(pid, page_type, screenshot_dir)
                if path and is_first_of_group:
                    screenshot_images.append((row_idx, ci, path))
                    v = ""
                else:
                    v = "无" if not path else ""
            else:
                v = r.get(k)
                v = format_export_value(k, v)
            if isinstance(v, (dict, list)):
                v = str(v)
            ws.cell(row_idx, ci, v)
        row_idx += 1

    _img_w, _img_h = 80, 50
    for r, c, path in screenshot_images:
        add_screenshot_image(ws, r, c, path, width=_img_w, height=_img_h)
    for r, c, _ in screenshot_images:
        ws.row_dimensions[r].height = _img_h * 0.75
    for c in {c for _, c, _ in screenshot_images}:
        ws.column_dimensions[get_column_letter(c)].width = max(_img_w / 7, 10)

    for grp in promotion_groups:
        if grp["count"] <= 1:
            continue
        start_row, end_row = grp["start"], grp["start"] + grp["count"] - 1
        for ci, k in enumerate(cols, 1):
            if k in ten_min_cols:
                continue
            ws.merge_cells(start_row=start_row, start_column=ci, end_row=end_row, end_column=ci)
            cell = ws.cell(start_row, ci)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fn, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/accounts")
def api_accounts():
    """获取账号列表"""
    return jsonify(list_accounts())


@app.route("/api/accounts/<account_id>", methods=["DELETE"])
def api_delete_account(account_id: str):
    """删除账号"""
    if delete_account(account_id):
        return jsonify({"ok": True})
    return jsonify({"error": "账号不存在"}), 404


@app.route("/api/request_add_account", methods=["POST"])
def api_request_add_account():
    """请求添加账号（写入 pending action，由 main 轮询执行）"""
    write_pending_action("add_account")
    return jsonify({"ok": True})


@app.route("/api/request_open_browser", methods=["POST"])
def api_request_open_browser():
    """请求打开账号浏览器（写入 pending action）"""
    data = request.get_json() or {}
    write_pending_action("open_browser", account_id=data.get("account_id", ""))
    return jsonify({"ok": True})


@app.route("/api/request_collect", methods=["POST"])
def api_request_collect():
    """请求采集（写入 pending action）"""
    data = request.get_json() or {}
    write_pending_action(
        "collect",
        account_id=data.get("account_id", ""),
        start_time=data.get("start_time") or "",
        end_time=data.get("end_time") or "",
    )
    return jsonify({"ok": True})


@app.route("/api/config", methods=["GET", "POST"])
def api_config():
    """获取/保存 UI 配置（模式、分辨率）"""
    path = Path(UI_CONFIG_FILE)
    path.parent.mkdir(parents=True, exist_ok=True)
    if request.method == "POST":
        data = request.get_json() or {}
        cfg = {**DEFAULT_UI_CONFIG, **data}
        path.write_text(json.dumps(cfg, ensure_ascii=False, indent=2))
        return jsonify(cfg)
    if path.exists():
        cfg = json.loads(path.read_text(encoding="utf-8"))
        return jsonify({**DEFAULT_UI_CONFIG, **cfg})
    return jsonify(DEFAULT_UI_CONFIG)


@app.route("/api/clear_all_data", methods=["POST"])
def api_clear_all_data():
    """一键删除所有订单数据和截图（保留账号）"""
    import shutil
    count = clear_all_data()
    screenshot_path = Path(SCREENSHOT_DIR)
    if screenshot_path.exists():
        shutil.rmtree(screenshot_path, ignore_errors=True)
    return jsonify({"ok": True, "deleted_orders": count})


@app.route("/api/logs")
def api_logs():
    """获取最近日志（开发模式用）"""
    lines = int(request.args.get("lines", 100))
    if not Path(LOG_FILE).exists():
        return jsonify({"lines": []})
    with open(LOG_FILE, "r", encoding="utf-8", errors="replace") as f:
        all_lines = f.readlines()
    return jsonify({"lines": all_lines[-lines:]})


def run_app(host: str = "0.0.0.0", port: int = 5000):
    Path("templates").mkdir(exist_ok=True)
    app.run(host=host, port=port, debug=False, use_reloader=False)


if __name__ == "__main__":
    run_app()
